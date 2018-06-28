# encoding:utf-8

from bs4 import BeautifulSoup
import requests
import json
import re
import openpyxl
import xml.etree.ElementTree as ET

headers = {
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/66.0.3359.139 Safari/537.36"
}

def get_school_id(school_name):
    Referer = "https://gkcx.eol.cn/soudaxue/queryschool.html?&keyWord1={}".format(school_name)
    data_url = "https://data-gkcx.eol.cn/soudaxue/queryschool.html"
    params = {
        "messtype" : "jsonp",
        "_":"1530074932531",
        "callback" : "",
        "keyWord1" : school_name
    }
    headers["Referer"] = Referer.encode('utf-8')

    response = requests.request("GET", data_url,headers=headers,params=params)
    # 返回数据包含 ();，需要特殊处理
    text = ((response.text).split(');',1)[0]).split('(',1)[1]
    j = json.loads(text)
    school_list = j['school']
    return [{'schoolid': i['schoolid'], 'schoolname':i['schoolname']} for i in school_list]

def get_xml(school):
    id = school['schoolid']
    Referer = "https://gkcx.eol.cn/schoolhtm/schoolTemple/school{}.htm".format(id)
    headers["Referer"] = Referer.encode('utf-8')
    data_url = "https://gkcx.eol.cn/commonXML/schoolSpecialPoint/schoolSpecialPoint{}_10014_10035.xml".format(id)
    response = requests.request("GET", data_url, headers=headers)
    response.encoding='utf-8'
    return response.text

def gen_excel(school,xml,wb):
    sheet = wb.create_sheet(title='各专业历年录取分数线')
    # 设置专业列的列宽
    sheet.column_dimensions['B'].width = 40
    sheet.freeze_panes = 'A1'
    sheet['A1'] = '年份'
    sheet['B1'] = '专业'
    sheet['C1'] = '最高分'
    sheet['D1'] = '平均分'
    sheet['E1'] = '最低分'
    sheet['F1'] = '批次'
    sheet['G1'] = '录取批次'

    areapionts = ET.fromstring(xml)
    column = 1 
    for areapiont in areapionts:
        column += 1
        sheet.cell(column,1).value = areapiont.find('year').text
        sheet.cell(column,2).value = areapiont.find('specialname').text
        sheet.cell(column,3).value = areapiont.find('maxfs').text
        sheet.cell(column,4).value = areapiont.find('varfs').text
        sheet.cell(column,5).value = areapiont.find('minfs').text
        sheet.cell(column,6).value = areapiont.find('pc').text
        sheet.cell(column,7).value = areapiont.find('stype').text
    wb.save('{}.xlsx'.format(school['schoolname']))

if __name__ == "__main__":
    school_name = input("Please the school name：")
    # school_name = '南京邮电大学'
    school_infos = get_school_id(school_name)
    print("共检索到 {} 个高校：{}".format(len(school_infos), [info['schoolname'] for info in school_infos]))
    for school in school_infos:
        wb = None
        try:
            wb = openpyxl.load_workbook('{}.xlsx'.format(school['schoolname']))
        except FileNotFoundError:
            wb = openpyxl.Workbook()
        # 删除之前的 sheet 页
        for sheet in wb.sheetnames:
            wb.remove(wb[sheet])
        xml = get_xml(school)
        gen_excel(school,xml,wb)
    print("数据获取完成，已下载到脚本目录")
    